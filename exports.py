"""Export utilities for PDF and Excel formats."""

from io import BytesIO
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle


def export_ledger_to_excel(party, transactions, total_receivable, total_payable, net_balance):
    """Export ledger data to Excel format."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ledger"
    
    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # Title
    ws.merge_cells("A1:F1")
    title = ws["A1"]
    title.value = f"Ledger Report - {party.name}"
    title.font = Font(bold=True, size=14)
    title.alignment = center_alignment
    
    # Summary section
    ws.merge_cells("A3:B3")
    ws["A3"] = "Total Receivable:"
    ws["A3"].font = Font(bold=True)
    ws["C3"] = total_receivable
    ws["C3"].number_format = '₹ #,##0.00'
    
    ws.merge_cells("A4:B4")
    ws["A4"] = "Total Payable:"
    ws["A4"].font = Font(bold=True)
    ws["C4"] = total_payable
    ws["C4"].number_format = '₹ #,##0.00'
    
    ws.merge_cells("A5:B5")
    ws["A5"] = "Net Balance:"
    ws["A5"].font = Font(bold=True)
    ws["C5"] = net_balance
    ws["C5"].number_format = '₹ #,##0.00'
    
    # Transactions table headers
    headers = ["Date", "Type", "Description", "Debit (₹)", "Credit (₹)", "Balance (₹)"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
    
    # Data rows
    running_balance = 0
    for row_num, tx in enumerate(transactions, 8):
        ws[f"A{row_num}"] = tx.date
        ws[f"A{row_num}"].number_format = "dd-mmm-yyyy"
        
        ws[f"B{row_num}"] = tx.transaction_type.replace("_", " ").title()
        ws[f"C{row_num}"] = tx.description or ""
        
        # Debit/Credit logic
        if tx.transaction_type in ["udhari_given", "repayment_received"]:
            ws[f"D{row_num}"] = tx.amount
            running_balance += tx.amount
        elif tx.transaction_type in ["cash_in"]:
            ws[f"D{row_num}"] = tx.amount
            running_balance += tx.amount
        else:
            ws[f"E{row_num}"] = tx.amount
            running_balance -= tx.amount
        
        ws[f"F{row_num}"] = running_balance
        
        # Format numbers
        ws[f"D{row_num}"].number_format = '₹ #,##0.00'
        ws[f"E{row_num}"].number_format = '₹ #,##0.00'
        ws[f"F{row_num}"].number_format = '₹ #,##0.00'
        
        # Borders
        for col in ["A", "B", "C", "D", "E", "F"]:
            ws[f"{col}{row_num}"].border = border
    
    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    
    # Add footer with export date
    footer_row = len(transactions) + 9
    ws.merge_cells(f"A{footer_row}:F{footer_row}")
    footer = ws[f"A{footer_row}"]
    footer.value = f"Exported on {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}"
    footer.font = Font(italic=True, size=10, color="666666")
    
    return wb


def export_transactions_to_excel(transactions, transaction_types):
    """Export all transactions to Excel format."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"
    
    # Header styling
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # Title
    ws.merge_cells("A1:F1")
    title = ws["A1"]
    title.value = "Transaction Report"
    title.font = Font(bold=True, size=14)
    title.alignment = center_alignment
    
    # Headers
    headers = ["Date", "Type", "Party", "Description", "Amount (₹)", "Created"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
    
    # Data rows
    for row_num, tx in enumerate(transactions, 4):
        ws[f"A{row_num}"] = tx.date
        ws[f"A{row_num}"].number_format = "dd-mmm-yyyy"
        
        ws[f"B{row_num}"] = transaction_types.get(tx.transaction_type, tx.transaction_type)
        ws[f"C{row_num}"] = tx.party.name if tx.party else "N/A"
        ws[f"D{row_num}"] = tx.description or ""
        ws[f"E{row_num}"] = tx.amount
        ws[f"E{row_num}"].number_format = '₹ #,##0.00'
        
        ws[f"F{row_num}"] = tx.created_at
        ws[f"F{row_num}"].number_format = "dd-mmm-yyyy hh:mm"
        
        # Borders and alignment
        for col in ["A", "B", "C", "D", "E", "F"]:
            ws[f"{col}{row_num}"].border = border
            if col in ["E"]:
                ws[f"{col}{row_num}"].alignment = Alignment(horizontal="right")
    
    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 18
    
    # Footer
    footer_row = len(transactions) + 5
    ws.merge_cells(f"A{footer_row}:F{footer_row}")
    footer = ws[f"A{footer_row}"]
    footer.value = f"Exported on {datetime.now().strftime('%d-%m-%Y %H:%M:%S')} | Total Transactions: {len(transactions)}"
    footer.font = Font(italic=True, size=10, color="666666")
    
    return wb


def export_dashboard_summary_to_excel(total_cash_in, total_cash_out, total_receivable, total_payable, recent_transactions, transaction_types):
    """Export dashboard summary to Excel format."""
    wb = Workbook()
    
    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Title
    ws_summary.merge_cells("A1:B1")
    title = ws_summary["A1"]
    title.value = "Dashboard Summary"
    title.font = Font(bold=True, size=14)
    
    # Summary cards
    ws_summary["A3"] = "Metric"
    ws_summary["B3"] = "Amount (₹)"
    for cell in ["A3", "B3"]:
        ws_summary[cell].fill = header_fill
        ws_summary[cell].font = header_font
    
    metrics = [
        ("Total Cash In", total_cash_in),
        ("Total Cash Out", total_cash_out),
        ("Total Receivable", total_receivable),
        ("Total Payable", total_payable),
        ("Net Balance", total_cash_in - total_cash_out),
    ]
    
    for idx, (label, value) in enumerate(metrics, 4):
        ws_summary[f"A{idx}"] = label
        ws_summary[f"B{idx}"] = value
        ws_summary[f"B{idx}"].number_format = '₹ #,##0.00'
    
    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 16
    
    # Recent transactions sheet
    ws_recent = wb.create_sheet("Recent Transactions")
    headers = ["Date", "Type", "Party", "Description", "Amount (₹)"]
    for col_num, header in enumerate(headers, 1):
        cell = ws_recent.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
    
    for row_num, tx in enumerate(recent_transactions, 2):
        ws_recent[f"A{row_num}"] = tx.date
        ws_recent[f"A{row_num}"].number_format = "dd-mmm-yyyy"
        ws_recent[f"B{row_num}"] = transaction_types.get(tx.transaction_type, tx.transaction_type)
        ws_recent[f"C{row_num}"] = tx.party.name if tx.party else "N/A"
        ws_recent[f"D{row_num}"] = tx.description or ""
        ws_recent[f"E{row_num}"] = tx.amount
        ws_recent[f"E{row_num}"].number_format = '₹ #,##0.00'
    
    for col in ["A", "B", "C", "D", "E"]:
        ws_recent.column_dimensions[col].width = 18
    
    return wb


def export_ledger_to_pdf(party, transactions, total_receivable, total_payable, net_balance):
    """Export ledger data to PDF format."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#366092'),
        spaceAfter=12,
        alignment=1  # center
    )
    
    elements = []
    
    # Title
    title = Paragraph(f"Ledger Report - {party.name}", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Summary table
    summary_data = [
        ["Total Receivable", f"₹ {total_receivable:,.2f}"],
        ["Total Payable", f"₹ {total_payable:,.2f}"],
        ["Net Balance", f"₹ {net_balance:,.2f}"],
    ]
    
    summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, 2), colors.HexColor('#E8F0F8')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, 2), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Transactions table
    table_data = [["Date", "Type", "Description", "Debit (₹)", "Credit (₹)", "Balance (₹)"]]
    
    running_balance = 0
    for tx in transactions:
        debit = ""
        credit = ""
        
        if tx.transaction_type in ["udhari_given", "repayment_received", "cash_in"]:
            debit = f"{tx.amount:,.2f}"
            running_balance += tx.amount
        else:
            credit = f"{tx.amount:,.2f}"
            running_balance -= tx.amount
        
        table_data.append([
            tx.date.strftime("%d-%m-%Y"),
            tx.transaction_type.replace("_", " ").title(),
            tx.description or "",
            debit,
            credit,
            f"{running_balance:,.2f}",
        ])
    
    transactions_table = Table(table_data, colWidths=[0.9*inch, 1.2*inch, 1.5*inch, 1*inch, 1*inch, 1*inch])
    transactions_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (2, 1), (2, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
    ]))
    elements.append(transactions_table)
    
    # Footer
    elements.append(Spacer(1, 0.2*inch))
    footer_text = f"<i>Exported on {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}</i>"
    footer = Paragraph(footer_text, styles['Normal'])
    elements.append(footer)
    
    doc.build(elements)
    buffer.seek(0)
    return buffer


def export_transactions_to_pdf(transactions, transaction_types):
    """Export transactions to PDF format."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#366092'),
        spaceAfter=12,
        alignment=1
    )
    
    elements = []
    
    # Title
    title = Paragraph("Transaction Report", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # Transactions table
    table_data = [["Date", "Type", "Party", "Description", "Amount (₹)"]]
    
    for tx in transactions:
        table_data.append([
            tx.date.strftime("%d-%m-%Y"),
            transaction_types.get(tx.transaction_type, tx.transaction_type),
            tx.party.name if tx.party else "N/A",
            tx.description or "",
            f"₹ {tx.amount:,.2f}",
        ])
    
    transactions_table = Table(table_data, colWidths=[1*inch, 1.3*inch, 1.3*inch, 1.5*inch, 1*inch])
    transactions_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (3, 1), (3, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
    ]))
    elements.append(transactions_table)
    
    # Footer
    elements.append(Spacer(1, 0.2*inch))
    footer_text = f"<i>Exported on {datetime.now().strftime('%d-%m-%Y %H:%M:%S')} | Total: {len(transactions)} transactions</i>"
    footer = Paragraph(footer_text, styles['Normal'])
    elements.append(footer)
    
    doc.build(elements)
    buffer.seek(0)
    return buffer
